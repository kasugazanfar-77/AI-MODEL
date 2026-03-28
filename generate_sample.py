"""
ArthSetu AI - Sample Bank Statement Generator
----------------------------------------------
Generates a realistic Indian bank statement Excel file for testing.
Run this script first to create 'sample_bank_statement.xlsx'

Usage:
    python generate_sample.py
"""

import pandas as pd
import numpy as np
import random
from datetime import datetime, timedelta
import os


# ── Seed for reproducibility ──────────────────────────────────────────────────
random.seed(42)
np.random.seed(42)

# ── Sample transaction templates ──────────────────────────────────────────────
EXPENSE_TEMPLATES = {
    "Food & Dining": [
        ("SWIGGY ORDER PAYMENT", (150, 800)),
        ("ZOMATO FOOD DELIVERY", (200, 900)),
        ("BIGBASKET GROCERY", (500, 3000)),
        ("DMART HYPERMARKET PURCHASE", (800, 4000)),
        ("MCDONALD INDIA PVT LTD", (200, 600)),
        ("STARBUCKS COFFEE PAYMENT", (300, 800)),
        ("BLINKIT QUICK DELIVERY", (200, 1500)),
        ("HALDIRAMS SNACKS UPI", (100, 500)),
        ("PIZZA HUT ORDER ONLINE", (400, 1200)),
        ("LOCAL KIRANA STORE UPI", (100, 800)),
        ("ZEPTO DARK STORE DELIVERY", (150, 1000)),
        ("ZOMATO FOOD ORDER", (250, 700)),
    ],
    "Transport": [
        ("OLA CAB BOOKING", (80, 400)),
        ("UBER INDIA RIDE", (100, 500)),
        ("INDIAN RAILWAYS IRCTC", (200, 2000)),
        ("PETROL PUMP FUEL BHARAT", (500, 3000)),
        ("RAPIDO BIKE TAXI PAYMENT", (30, 150)),
        ("MUMBAI METRO CARD RECHARGE", (200, 500)),
        ("FASTAG TOLL RECHARGE NHAI", (100, 500)),
        ("MAKEMYTRIP FLIGHT BOOKING", (2000, 12000)),
        ("REDBUS TICKET BOOKING", (300, 1500)),
    ],
    "Shopping": [
        ("AMAZON INDIA PURCHASE", (300, 8000)),
        ("FLIPKART ORDER PAYMENT", (500, 6000)),
        ("MYNTRA FASHION ORDER", (400, 3000)),
        ("AJIO CLOTHING PURCHASE", (300, 2500)),
        ("NYKAA BEAUTY PRODUCTS", (200, 1500)),
        ("RELIANCE DIGITAL STORE", (1000, 20000)),
        ("DECATHLON SPORTS EQUIPMENT", (500, 5000)),
        ("CROMA ELECTRONICS", (1500, 25000)),
        ("MEESHO SOCIAL COMMERCE", (200, 1500)),
    ],
    "Bills & Utilities": [
        ("BESCOM ELECTRICITY BILL", (800, 4000)),
        ("AIRTEL FIBER BROADBAND BILL", (700, 1200)),
        ("JIO FIBER PAYMENT", (600, 1100)),
        ("AIRTEL MOBILE RECHARGE", (200, 600)),
        ("JIO PREPAID RECHARGE", (150, 500)),
        ("MAHANAGAR GAS LTD BILL", (400, 1500)),
        ("TATA SKY DTH RECHARGE", (200, 700)),
        ("BBMP PROPERTY TAX PAYMENT", (1000, 8000)),
        ("SOCIETY MAINTENANCE FEE", (1500, 6000)),
        ("WATER BILL MUNICIPAL CORP", (200, 600)),
    ],
    "Entertainment": [
        ("NETFLIX INDIA SUBSCRIPTION", (149, 649)),
        ("AMAZON PRIME VIDEO RENEWAL", (299, 999)),
        ("DISNEY PLUS HOTSTAR PLAN", (299, 899)),
        ("SPOTIFY PREMIUM MUSIC", (119, 189)),
        ("BOOKMYSHOW MOVIE TICKET", (150, 800)),
        ("PVR CINEMAS TICKET BOOKING", (200, 1000)),
        ("YOUTUBE PREMIUM INDIA", (139, 189)),
        ("DREAM11 FANTASY SPORTS", (50, 500)),
        ("SONY LIV SUBSCRIPTION", (299, 899)),
    ],
    "Healthcare": [
        ("APOLLO PHARMACY MEDICINE", (200, 2000)),
        ("1MG ONLINE PHARMACY ORDER", (150, 1500)),
        ("PRACTO DOCTOR CONSULTATION", (200, 1500)),
        ("MANIPAL HOSPITAL PAYMENT", (500, 15000)),
        ("THYROCARE LAB TEST PAYMENT", (300, 2000)),
        ("CULT FIT MEMBERSHIP", (1000, 3000)),
        ("GOLD GYM MONTHLY FEE", (800, 2500)),
        ("PHARMEASY MEDICINE DELIVERY", (100, 1200)),
        ("YOGA CLASS FEES", (500, 2000)),
    ],
    "Education": [
        ("UDEMY ONLINE COURSE", (499, 3999)),
        ("COURSERA SUBSCRIPTION", (1500, 4000)),
        ("UNACADEMY SUBSCRIPTION", (500, 3000)),
        ("BYJU LEARNING APP FEES", (2000, 8000)),
        ("SCHOOL FEES PAYMENT", (5000, 25000)),
        ("COACHING CLASS FEES", (2000, 10000)),
        ("SIMPLILEARN CERTIFICATION", (1000, 5000)),
    ],
    "EMI & Loan": [
        ("HOME LOAN EMI HDFC BANK", (10000, 40000)),
        ("PERSONAL LOAN EMI ICICI", (3000, 15000)),
        ("CAR LOAN EMI SBI", (5000, 20000)),
        ("CONSUMER DURABLE EMI BAJAJ", (1000, 5000)),
    ],
    "Insurance": [
        ("HEALTH INSURANCE PREMIUM STAR", (3000, 12000)),
        ("LIC LIFE INSURANCE PREMIUM", (2000, 10000)),
        ("CAR INSURANCE RENEWAL BAJAJ", (5000, 20000)),
        ("TERM PLAN PREMIUM HDFC LIFE", (1000, 8000)),
    ],
    "Shopping (ATM)": [
        ("ATM CASH WITHDRAWAL", (1000, 10000)),
    ],
    "Investment": [
        ("ZERODHA SIP INSTALLMENT", (500, 10000)),
        ("GROWW MUTUAL FUND SIP", (500, 5000)),
        ("PPFD DEPOSIT AXIS BANK", (1000, 25000)),
        ("LIC ULIP INVESTMENT PLAN", (2000, 10000)),
        ("UPSTOX EQUITY PURCHASE", (1000, 50000)),
    ],
}

INCOME_TEMPLATES = [
    "SALARY CREDIT NEFT {employer}",
    "MONTHLY SALARY DEPOSITED",
    "NEFT PAYROLL CREDIT {employer}",
    "CONSULTING FEE RECEIVED NEFT",
    "FREELANCE PROJECT PAYMENT",
    "DIVIDEND INCOME CREDITED",
    "INTEREST CREDITED FD SAVINGS",
    "REFUND CREDITED AMAZON",
    "CASHBACK REWARD CREDITED",
    "RENTAL INCOME PROPERTY",
]

EMPLOYERS = [
    "INFOSYS LTD", "WIPRO TECHNOLOGIES", "TCS TATA CONSULTANCY",
    "HDFC BANK", "ICICI BANK", "RELIANCE INDUSTRIES",
    "GOOGLE INDIA PVT LTD", "AMAZON INDIA PVTLTD",
    "STARTUP PVT LTD", "MNC INDIA OPERATIONS",
]


def generate_sample_statement(
    months: int = 4,
    monthly_income: float = 85000,
    output_path: str = "sample_bank_statement.xlsx",
) -> pd.DataFrame:
    """
    Generate a realistic bank statement with income and expenses.

    Args:
        months: Number of months to simulate
        monthly_income: Approximate monthly salary
        output_path: Excel file to create

    Returns:
        DataFrame of generated transactions
    """
    print(f"\n📝 Generating sample bank statement ({months} months, ₹{monthly_income:,.0f}/month income)...")

    transactions = []
    start_date = datetime(2024, 1, 1)
    employer = random.choice(EMPLOYERS)

    for month_offset in range(months):
        month_start = start_date + timedelta(days=30 * month_offset)

        # ── Salary credit (1st or 2nd working day) ───────────────────────────
        salary_day = random.randint(1, 5)
        salary_date = month_start + timedelta(days=salary_day)
        salary_amount = monthly_income * random.uniform(0.97, 1.03)

        desc = random.choice(INCOME_TEMPLATES).format(employer=employer)
        transactions.append({
            "Date": salary_date,
            "Description": desc,
            "Amount": round(salary_amount, 2),
        })

        # ── Occasional extra income ──────────────────────────────────────────
        if random.random() < 0.3:
            extra_date = month_start + timedelta(days=random.randint(10, 25))
            transactions.append({
                "Date": extra_date,
                "Description": "CONSULTING FEE RECEIVED NEFT",
                "Amount": round(random.uniform(5000, 20000), 2),
            })

        # ── EMI payments (fixed monthly) ─────────────────────────────────────
        emi_date = month_start + timedelta(days=random.randint(3, 7))
        transactions.append({
            "Date": emi_date,
            "Description": "HOME LOAN EMI HDFC BANK AUTO DEBIT",
            "Amount": -round(random.uniform(18000, 22000), 2),
        })

        if random.random() < 0.6:
            transactions.append({
                "Date": emi_date + timedelta(days=1),
                "Description": "CAR LOAN EMI SBI AUTO DEBIT",
                "Amount": -round(random.uniform(7000, 9000), 2),
            })

        # ── Investment SIP ────────────────────────────────────────────────────
        if random.random() < 0.8:
            sip_date = month_start + timedelta(days=random.randint(5, 10))
            transactions.append({
                "Date": sip_date,
                "Description": "ZERODHA COIN SIP MUTUAL FUND",
                "Amount": -round(random.uniform(3000, 8000), 2),
            })

        # ── Daily expenses ────────────────────────────────────────────────────
        n_expenses = random.randint(18, 35)
        for _ in range(n_expenses):
            # Pick random category (weighted)
            categories = list(EXPENSE_TEMPLATES.keys())
            weights = [15, 8, 12, 10, 8, 5, 3, 0, 0, 3, 6]  # match category order
            if len(weights) < len(categories):
                weights.extend([3] * (len(categories) - len(weights)))

            cat = random.choices(categories, weights=weights[:len(categories)], k=1)[0]
            template, (amt_min, amt_max) = random.choice(EXPENSE_TEMPLATES[cat])

            txn_date = month_start + timedelta(days=random.randint(0, 28))
            amount = -round(random.uniform(amt_min, amt_max), 2)

            transactions.append({
                "Date": txn_date,
                "Description": template,
                "Amount": amount,
            })

        # ── Insurance (quarterly) ─────────────────────────────────────────────
        if month_offset % 3 == 0:
            ins_date = month_start + timedelta(days=random.randint(10, 20))
            transactions.append({
                "Date": ins_date,
                "Description": "HEALTH INSURANCE PREMIUM STAR HEALTH",
                "Amount": -round(random.uniform(3500, 6000), 2),
            })

        # ── Society maintenance ───────────────────────────────────────────────
        maint_date = month_start + timedelta(days=random.randint(1, 10))
        transactions.append({
            "Date": maint_date,
            "Description": "SOCIETY MAINTENANCE CHARGES NEFT",
            "Amount": -round(random.uniform(2000, 4000), 2),
        })

    # Build DataFrame
    df = pd.DataFrame(transactions)
    df = df.sort_values("Date").reset_index(drop=True)

    # ── Write to Excel ────────────────────────────────────────────────────────
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Bank Statement")

        # Style the Excel sheet
        ws = writer.sheets["Bank Statement"]

        # Column widths
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 15

        # Header styling
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
        header_font = Font(color="6C63FF", bold=True, size=12)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    print(f"   ✅ Created: {output_path}")
    print(f"   📊 Total transactions: {len(df)}")
    print(f"   💰 Total Income  : ₹{df[df['Amount'] > 0]['Amount'].sum():,.2f}")
    print(f"   💸 Total Expense : ₹{df[df['Amount'] < 0]['Amount'].abs().sum():,.2f}")
    print(f"   📅 Date range    : {df['Date'].min().date()} to {df['Date'].max().date()}")

    return df


if __name__ == "__main__":
    generate_sample_statement(months=4, monthly_income=85000)
    print("\n✅ Sample file ready: sample_bank_statement.xlsx")
    print("   Now run: python app.py sample_bank_statement.xlsx")
